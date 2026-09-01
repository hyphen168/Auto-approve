"""Excel 表格处理：AI 数据分析、分析报告写入、按表头生成数据、一键图表。"""
import io
import json
import re

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from . import llm

EXCEL_SYSTEM = "你是数据分析专家，擅长从表格数据中提炼结论、发现问题并给出可执行的建议。"


def _load(content, data_only=True):
    return load_workbook(io.BytesIO(content), data_only=data_only)


def load_sheet_info(content):
    """读取工作簿每个工作表的前 20 行样例数据。"""
    wb = _load(content)
    info = {}
    sheets = wb.sheetnames
    for ws in wb.worksheets:
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20:
                break
            cleaned = ["" if v is None else str(v) for v in row]
            if any(str(v).strip() for v in cleaned):
                rows.append(cleaned)
        info[ws.title] = rows
    wb.close()
    return info, sheets


def _describe(info):
    lines = []
    for sheet, rows in info.items():
        lines.append("表名：%s" % sheet)
        lines.append("前 %d 行数据：" % len(rows))
        for r in rows:
            lines.append(" | ".join(r))
        lines.append("")
    return "\n".join(lines)


def analyze_sheet(cfg, content, extra=""):
    """AI 分析整个工作簿。"""
    info, _ = load_sheet_info(content)
    user = (
        "请分析这份 Excel 数据，输出以下内容：\n"
        "1) 总体说明（每个工作表的作用与字段含义）；\n"
        "2) 发现的问题（数据质量、异常值、缺失、格式等）；\n"
        "3) 业务分析与建议。\n"
        "用「-」列表输出。\n\n【数据】\n" + _describe(info) +
        (("\n【附加要求】" + extra) if extra else "")
    )
    return llm.chat(
        cfg,
        [{"role": "system", "content": EXCEL_SYSTEM},
         {"role": "user", "content": user}],
        max_tokens=4096,
        timeout=900,
    )


def add_analysis_sheet(content, analysis_text):
    """把分析文本写入名为「AI分析报告」的新工作表，返回新工作簿字节。"""
    wb = load_workbook(io.BytesIO(content), data_only=False)
    ws = wb.create_sheet("AI分析报告")
    for line in analysis_text.splitlines():
        ws.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _extract_json(text):
    """从 AI 输出中尽力提取 JSON 数组。"""
    text = text.strip()
    s, e = text.find("["), text.rfind("]")
    if s != -1 and e > s:
        text = text[s:e + 1]
    try:
        data = json.loads(text)
        return data if isinstance(data, list) else [data]
    except Exception:
        m = re.search(r"\{.*\}", text, re.S)
        if m:
            try:
                return [json.loads(m.group(0))]
            except Exception:
                return None
        return None


def generate_rows(cfg, content, sheet_name, n_rows, extra=""):
    """按第一行表头信息让 AI 生成 n 条数据，追加到工作表，返回(新工作簿字节, 起始行, 生成条数)。"""
    wb = _load(content)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError("找不到工作表「%s」，请在打开文件后选择正确的工作表。" % sheet_name)
    ws = wb[sheet_name]
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    if not any(headers):
        raise RuntimeError("该工作表第一行没有表头，请先填写表头再生成数据。")
    n_rows = max(1, int(n_rows))

    user = (
        "请根据表头生成 %d 条真实、合理且互不重复的数据。\n"
        "表头：%s\n"
        "附加要求：%s\n"
        "严格按如下格式只输出一个 JSON 数组（每个元素是对象，key 与表头完全一致），不要输出任何其他内容：\n"
        '[{"列1": 值, "列2": 值}, ...]'
        % (n_rows, headers, extra or "无")
    )
    ai_text = llm.chat(
        cfg,
        [{"role": "system", "content": EXCEL_SYSTEM},
         {"role": "user", "content": user}],
        max_tokens=8192,
        timeout=900,
    )
    data = _extract_json(ai_text)
    if not isinstance(data, list) or not data:
        raise RuntimeError("AI 返回的数据无法解析为 JSON 数组，请重试或更换模型。")

    wbw = load_workbook(io.BytesIO(content), data_only=False)
    wsw = wbw[sheet_name]
    start = wsw.max_row + 1
    for item in data[:n_rows]:
        if not isinstance(item, dict):
            item = {headers[0]: str(item)}
        wsw.append([item.get(h, "") for h in headers])
    buf = io.BytesIO()
    wbw.save(buf)
    wbw.close()
    return buf.getvalue(), start, len(data)


def add_chart(content, sheet_name, chart_type="bar"):
    """为数值列自动生成柱状图/折线图，返回新工作簿字节。"""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError("找不到工作表「%s」。" % sheet_name)
    ws = wb[sheet_name]
    max_col = ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

    numeric_cols = []
    for c in range(1, max_col + 1):
        vals = [ws.cell(row=r, column=c).value for r in range(2, ws.max_row + 1)]
        if any(isinstance(v, (int, float)) for v in vals):
            numeric_cols.append(c)
    if not numeric_cols:
        raise RuntimeError("工作表「%s」中没有数值列，无法生成图表。" % sheet_name)

    cat_name = str(headers[0] or "类别")
    chart = BarChart() if chart_type == "bar" else LineChart()
    chart.title = "%s 数据图" % sheet_name
    chart.x_axis.title = cat_name
    chart.y_axis.title = "数值"
    for c in numeric_cols[:5]:
        chart.add_data(
            Reference(ws, min_col=c, min_row=1, max_row=ws.max_row),
            titles_from_data=True,
        )
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=ws.max_row))
    chart.width = 24
    chart.height = 12

    ws.add_chart(chart, "%s2" % get_column_letter(ws.max_column + 2))
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()